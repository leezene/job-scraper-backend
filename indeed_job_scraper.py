from bs4 import BeautifulSoup
import requests
from random import random
from time import sleep
from email.message import EmailMessage
from collections import namedtuple
import smtplib
import csv
from openpyxl import load_workbook

EmailCredentials = namedtuple("EmailCredentials", ['username', 'password', 'sender', 'recipient'])


def generate_url(domain, date_posted, job_title, job_location):
    url_template = "https://" + domain + "/jobs?q={}&l={}&fromage={}"
    print(job_title)
    url = url_template.format(job_title, job_location, date_posted)
    return url


def save_record_to_csv(record, filepath, create_new_file=False):
    """Save an individual record to file; set `new_file` flag to `True` to generate new file"""
    header = ["JobTitle", "Company", "Location", "Summary", "PostDate", "JobUrl"]
    if create_new_file:
        wb = load_workbook(filename=filepath)
        wb.remove(wb.worksheets[0])
        wb.create_sheet()
        ws = wb.worksheets[0]
        ws.append(header)
        wb.save(filepath)
    else:
        wb = load_workbook(filename=filepath)
        # Select First Worksheet
        ws = wb.worksheets[0]
        ws.append(record)
        wb.save(filepath)


def collect_job_cards_from_page(html):
    soup = BeautifulSoup(html, 'html.parser')
    cards = soup.find_all('a', 'result')
    return cards, soup


def sleep_for_random_interval():
    seconds = random() * 10
    sleep(seconds)


def request_jobs_from_indeed(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.110 Safari/537.36",
        "Accept-Encoding": "gzip, deflate",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "en"
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.text
    else:
        return None


def find_next_page(soup):
    try:
        pagination = soup.find("a", {"aria-label": "Next"}).get("href")
        return "https://ae.indeed.com" + pagination
    except AttributeError:
        return None


def extract_job_card_data(card):
    atag = card.h2.find('span', {"title": True})
    try:
        job_title = atag.get('title') if atag.get('title') is not None else ''
    except AttributeError:
        job_title = ''
    try:
        company = card.find('span', 'companyName').a.text.strip() if card.find('span', 'companyName').a is not None \
            else card.find('span', 'companyName').text.strip()
    except AttributeError:
        company = ''
    try:
        location = card.find('div', 'companyLocation').text.strip()
    except AttributeError:
        location = ''
    try:
        job_summary = card.find('div', 'job-snippet').li.text.strip()
    except AttributeError:
        job_summary = ''
    try:
        post_date = card.find('span', 'date').text.strip()
    except AttributeError:
        post_date = ''
    try:
        salary = card.find('span', 'salarytext').text.strip()
    except AttributeError:
        salary = ''
    job_url = 'https://ae.indeed.com' + card.get('href')
    return job_title, company, location, job_summary, post_date, job_url


def main(domain, date_posted, job_title, job_location, filepath, email=None):
    unique_jobs = set()  # track job urls to avoid collecting duplicate records
    print("Starting to scrape indeed for `{}` in `{}`".format(job_title, job_location))
    url = generate_url(domain, date_posted, job_title, job_location)
    save_record_to_csv(None, filepath, create_new_file=True)
    page = 1

    while page < 3:
        print(url)
        html = request_jobs_from_indeed(url)
        if not html:
            break
        cards, soup = collect_job_cards_from_page(html)
        for card in cards:
            record = extract_job_card_data(card)
            raw_date = record[4]

            if 'Just posted' in raw_date or 'Today' in raw_date:
                date = 0
            else:
                date = int(raw_date.replace('days ago', '').replace('day ago', '').replace('+', ''))

            if date_posted == None:
                if not record[-1] in unique_jobs:
                    save_record_to_csv(record, filepath)
                    unique_jobs.add(record[-1])
            elif date <= int(date_posted) :
                if not record[-1] in unique_jobs:
                    save_record_to_csv(record, filepath)
                    unique_jobs.add(record[-1])

            else:
                pass
            print(record[4])

        sleep_for_random_interval()
        url = find_next_page(soup)
        page = page + 1
        if not url:
            break
    print('Finished collecting {:,d} job postings.'.format(len(unique_jobs)))
