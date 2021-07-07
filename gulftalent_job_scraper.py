from bs4 import BeautifulSoup
import requests
from random import random
from time import sleep
from email.message import EmailMessage
from collections import namedtuple
import smtplib
import csv
from openpyxl import load_workbook
import os
from datetime import datetime


def save_record_to_csv(record, filepath, create_new_file=False):
    """Save an individual record to file; set `new_file` flag to `True` to generate new file"""
    header = ["JobTitle", "Company", "Location", "Summary", "PostDate", "JobUrl"]
    if create_new_file:
        wb = load_workbook(filename=filepath)
        wb.remove(wb.worksheets[0])
        wb.create_sheet()
        ws = wb.worksheets[0]
        ws.append(header)
        wb.save(filepath)
    else:
        wb = load_workbook(filename=filepath)
        # Select First Worksheet
        ws = wb.worksheets[0]
        ws.append(record)
        wb.save(filepath)


def givemejson(title):
    url = 'https://www.gulftalent.com/api/jobs/search?condensed=false&config%5Bfilters%5D=ENABLED&config%5Bresults%5D=UNFILTERED&filters%5Bcountry%5D%5B0%5D=10111111000000&include_scraped=1&version=2'
    req = requests.get(url + '&limit=10&search_keyword=' + title )
    return req.json()['results']['data']


def main(domain, date_posted, job_title, job_location, filepath, email=None):
    job_base_url = 'https://www.gulftalent.com'
    save_record_to_csv(None, filepath, create_new_file=True)
    jobs = givemejson(job_title)
    for job in jobs:
        datetime_time = datetime.fromtimestamp(job['posted_date_ts'])
        d1 = datetime.now()
        delta = d1 - datetime_time
        result = job['title'], job['company_name'], job['location'], '', str(delta.days) + ' days ago', job_base_url + job['link']
        save_record_to_csv(result, filepath)
